from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl import Workbook

def magic(number):
    return int(''.join(str(i) for i in number))

r = open("asdfaf.txt","r")
liste= []
for i in r:
    listes = list()
    for j in i:
        if(j.isdigit()):
            j = int(j)
            listes.append(j)
    listes = magic(listes)
    liste.append(listes)




switch = " "
workbook_say = Workbook()
sheet_say = workbook_say.active

workbook_soz = Workbook()
sheet_soz = workbook_soz.active

workbook_ea = Workbook()
sheet_ea = workbook_ea.active

workbook_dil = Workbook()
sheet_dil = workbook_dil.active


driver = webdriver.Firefox()
i = 0 ;


for i in liste:
    puanlar = list()
    siralamalar = list()
    yil = list()
    tur = list()
    katsayi = list()
    obp = list()
    yerl = list()
    tyt_t = list()
    tyt_s = list()
    tyt_m = list()
    tyt_f = list()
    ayt_m = list()
    ayt_t = list()
    ayt_t1 = list()
    ayt_c1 = list()

    ayt_tde = list()
    ayt_tar1 = list()
    ayt_cog1 = list()
    ayt_tar2 = list()
    ayt_cog2 = list()
    ayt_fel = list()
    ayt_din = list()


    ayt_kim = list()
    ayt_fiz = list()
    ayt_biy = list()


    ydt_dil = list()
    driver.get("https://yokatlas.yok.gov.tr/netler-tablo.php?b={}".format(i))
    login = driver.find_element_by_xpath("""//*[@id="mydata"]""")
    puan_turu = driver.find_element_by_xpath("""/html/body/div/div[2]/div[2]/div[1]/div[1]/div/h2/strong""")
    source = BeautifulSoup(login.get_attribute("outerHTML"),"html.parser")
    orta = source.find_all("td",attrs={"class":"text-center"})#!/usr/bin/env python3


    if("EA" in puan_turu.text):
        switch = "ea"
        k = 1 ;
        for i in orta:
            if(k%13 == 1):
                yil.append(i.text)
                k= k +1
            elif(k%13 == 2):
                tur.append(i.text)
                k= k +1
            elif(k%13 == 3):
                katsayi.append(i.text)
                k= k +1
            elif(k%13 == 4):
                obp.append(i.text)
                k= k +1
            elif(k%13 == 5):
                yerl.append(i.text)
                k= k +1
            elif(k%13 == 6):
                tyt_t.append(i.text)
                k= k +1

            elif(k%13 == 7):
                tyt_s.append(i.text)
                k= k +1
            elif(k%13 == 8):
                tyt_m.append(i.text)
                k= k +1
            elif(k%13 == 9):
                tyt_f.append(i.text)
                k= k +1
            elif(k%13 == 10):
                ayt_m.append(i.text)
                k= k +1
            elif(k%13 == 11):
                ayt_t.append(i.text)
                k= k +1
            elif(k%13 == 12):
                ayt_t1.append(i.text)
                k= k +1
            elif(k%13 == 0):
                ayt_c1.append(i.text)
                k = 1



    elif("SAY" in puan_turu.text):
        switch = "say"
        k = 1 ;
        for i in orta:
            if(k%13 == 1):
                yil.append(i.text)
                k= k +1
            elif(k%13 == 2):
                tur.append(i.text)
                k= k +1
            elif(k%13 == 3):
                katsayi.append(i.text)
                k= k +1
            elif(k%13 == 4):
                obp.append(i.text)
                k= k +1
            elif(k%13 == 5):
                yerl.append(i.text)
                k= k +1
            elif(k%13 == 6):
                tyt_t.append(i.text)
                k= k +1

            elif(k%13 == 7):
                tyt_s.append(i.text)
                k= k +1
            elif(k%13 == 8):
                tyt_m.append(i.text)y
                k= k +1
            elif(k%13 == 9):
                tyt_f.append(i.text)
                k= k +1
            elif(k%13 == 10):
                ayt_m.append(i.text)
                k= k +1
            elif(k%13 == 11):
                ayt_fiz.append(i.text)
                k= k +1
            elif(k%13 == 12):
                ayt_kim.append(i.text)
                k= k +1
            elif(k%13 == 0):
                ayt_biy.append(i.text)
                k = 1



    elif("DİL" in puan_turu.text):
        switch = "dil"
        k = 1 ;
        for i in orta:
            if(k%10 == 1):
                yil.append(i.text)
                k= k +1
            elif(k%10 == 2):
                tur.append(i.text)
                k= k +1
            elif(k%10 == 3):
                katsayi.append(i.text)
                k= k +1
            elif(k%10 == 4):
                obp.append(i.text)
                k= k +1
            elif(k%10 == 5):
                yerl.append(i.text)
                k= k +1
            elif(k%10 == 6):
                tyt_t.append(i.text)
                k= k +1

            elif(k%10 == 7):
                tyt_s.append(i.text)
                k= k +1
            elif(k%10 == 8):
                tyt_m.append(i.text)
                k= k +1
            elif(k%10 == 9):
                tyt_f.append(i.text)
                k= k +1

            elif(k%10 == 0):
                ydt_dil.append(i.text)
                k = 1





    elif("SÖZ" in puan_turu.text):
        switch = "söz"
        k = 1 ;
        for i in orta:
            if(k%16 == 1):
                yil.append(i.text)
                k= k +1
                print(k)
            elif(k%16 == 2):
                tur.append(i.text)
                k= k +1
                print(k)
            elif(k%16 == 3):
                katsayi.append(i.text)
                k= k +1
                print(k)
            elif(k%16 == 4):
                obp.append(i.text)
                k= k +1
                print(k)
            elif(k%16 == 5):
                yerl.append(i.text)
                k= k +1
                print(k)
            elif(k%16 == 6):
                tyt_t.append(i.text)
                k= k +1
                print(k)

            elif(k%16 == 7):
                tyt_s.append(i.text)
                k= k +1
                print(k)
            elif(k%16 == 8):
                tyt_m.append(i.text)
                k= k +1
                print(k)
            elif(k%16 == 9):
                tyt_f.append(i.text)
                k= k +1
                print(k)
            elif(k%16 == 10):
                ayt_tde.append(i.text)
                k= k +1
                print(k)
            elif(k%16 == 11):
                ayt_tar1.append(i.text)
                k= k +1
                print(k)
            elif(k%16 == 12):
                ayt_cog1.append(i.text)
                k= k +1
                print(k)
            elif(k%16 == 13):
                ayt_tar2.append(i.text)
                k = 1+k
                print(k)
            elif(k%16 == 14):
                ayt_cog2.append(i.text)
                k= k +1
                print(k)
            elif(k%16 == 15):
                ayt_fel.append(i.text)
                k = 1+k
                print(k)
            elif(k%16 == 0):
                ayt_din.append(i.text)
                k= 1
                print(k)










    bas = source.find_all("a",href=True)
    for j in bas:
        driver.get("https://yokatlas.yok.gov.tr/"+j['href'])
        ogin = driver.find_element_by_xpath("""/html/body/div[2]/div[1]/div[4]/span/a[1]""")
        driver.implicitly_wait(10)
        ogin.click()
        puan=driver.find_element_by_xpath("""//*[@id="icerik_1000_1"]/table[3]/tbody/tr[1]/td[2]""")
        siralama = driver.find_element_by_xpath("""//*[@id="icerik_1000_1"]/table[3]/tbody/tr[3]/td[2]""")
        puanlar.append(puan.text)
        siralamalar.append(siralama.text)


    if(switch == "ea"):
        z = 0
        for i in puanlar:

            sheet_ea.append([bas[z]['href'],switch,puanlar[z],siralamalar[z],obp[z],tyt_t[z],tyt_s[z],tyt_m[z],tyt_f[z],ayt_m[z],ayt_t[z],ayt_t1[z],ayt_c1[z]])
            workbook_ea.save("ea.xlsx")
            z = z+1
    elif(switch == "say"):
        z = 0
        for i in puanlar:

            sheet_say.append([bas[z]['href'],switch,puanlar[z],siralamalar[z],obp[z],tyt_t[z],tyt_s[z],tyt_m[z],tyt_f[z],ayt_m[z],ayt_fiz[z],ayt_kim[z],ayt_biy[z]])
            workbook_say.save("say.xlsx")
            z = z+1
    elif(switch == "söz"):
        z = 0
        for i in tyt_t:

            sheet_soz.append([bas[z]['href'],switch,puanlar[z],siralamalar[z],obp[z],tyt_t[z],tyt_s[z],tyt_m[z],tyt_f[z],ayt_tde[z],ayt_tar1[z],ayt_cog1[z],ayt_tar2[z],ayt_cog2[z],ayt_fel[z],ayt_din[z]])
            workbook_soz.save("soz.xlsx")
            z = z+1
    elif(switch == "dil"):
        z = 0
        for i in puanlar:

            sheet_dil.append([bas[z]['href'],switch,puanlar[z],siralamalar[z],obp[z],tyt_t[z],tyt_s[z],tyt_m[z],tyt_f[z],ydt_dil[z]])
            workbook_dil.save("dil.xlsx")
            z = z+1


    # for i in tyt_t:
    #     print(i)
    # print(len(tyt_t))
    # print(len(yerl))
    # print(len(puanlar))





# passw=driver.find_element_by_xpath("""//*[@id="user_pass"]""")
# passw.send_keys("Serkaneren68.")
#
#
# buton = driver.find_element_by_xpath("""//*[@id="wp-submit"]""")
# buton.click()
# driver.get("https://sinavsayaci.com/wp-admin/post-new.php")
# elem = driver.find_element_by_xpath("""//*[@id="post-title-0"]""")
#
# elem.send_keys("pycons")
#
#
#
#
# katog = driver.find_element_by_xpath("""//*[@id="inspector-checkbox-control-7"]""")
# katog.click()
# sleep(10)
# driver.close()
